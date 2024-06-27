import datetime
from openpyxl import Workbook
import openpyxl
import sys
import os
import shutil
import hashlib
import re

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook

from cols import cols

# Purpose: Prepare HighStage album with photos for ingestion to Piwigo
# Pre requisites: Album.xlsx and Pic.xlsx is created from Highstage
# Licence: GNU 2.0
# Author: Ottar Kvindesland, 2024
# Reference: https://piwigo.miraheze.org/wiki/HighstageExport

class getId(cols):
    def __init__(self, pa):
        
        super().__init__()
        
        pwgfile = self.dbdumpdir + self.PiwigoPic
        self.pwg_wb = load_workbook(filename=pwgfile)
        self.wpwg = self.pwg_wb.worksheets[0]
        
        pwgafile = self.dbdumpdir + self.PiwigoAlbum
        self.pwa_wb = load_workbook(filename=pwgafile)
        self.wpwa = self.pwa_wb.worksheets[0]
        
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        if not (os.path.isfile(pwgfile) or os.path.islink(pwgfile)):
            print ('Missing File :', pwgfile)
        elif not (os.path.isfile(pwgafile) or os.path.islink(pwgafile)):
            print ('Missing File :', pwgafile)
        elif not (os.path.isfile(picfile) or os.path.islink(picfile)):
            print ('Missing File :', picfile)
        
        runPic = False
        runAlb = False
        runAlbImg = False
        if len(pa) >= 2:
            if pa[1] == 'P':
                runPic = True
            elif pa[1] == 'A':
                runAlb = True
            elif pa[1] == 'I':
                runAlbImg = True
        else:
            runPic = True
            runAlb = True
            runAlbImg = True
        
        if runPic:
            self.getSourcePicsId()
        if runAlb:
            self.getSourceAlbumId()
        if runAlbImg:
            self.defineBottomAlbumImages()
        
        if runPic:
            self.pic_wb.save(self.subdir + self.fOutputPic)
        if runAlb:
            self.album_wb.save(self.subdir + self.fOutputAlbum)
        
    def getSourcePicsId(self):
        
        n = 0
        tagged = 0
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n += 1
            if (n % 100) == 0:
                print ('Indexing ', n, ' pics')
            if n > 1:
                if pic[self.cpPath] is not None:
                    destpath = pic[self.cpPath].value
                    if destpath != '' and n>1:
                        pwgref = self.getPiwigoPicRef(destpath)
                        if pwgref.isnumeric():
                            self.wp.cell(row=n, column=self.cpPiwigoId+1).value = int(pwgref)
                            tagged += 1
        print ('\nReferenced ', tagged, ' photos\n')
    
    def getPiwigoPicRef(self, src):
        
        n = 0
        lastMatch = ''
        p_id = ''
        for ppic in self.wpwg.iter_rows(min_row=1, max_row=self.wpwg.max_row, min_col=1, max_col=self.pplastcol, values_only=False):
            if ppic[self.pppath] is not None:
                
                if str(ppic[self.pppath].value).endswith(str(src)):
                    n += 1
                    p_id = str(ppic[self.ppid].value)
                    if p_id == lastMatch:
                        print ('ERROR, Pic: source file used multiple times: ', src)
                    lastMatch = p_id
        return p_id

    def getSourceAlbumId(self):
        
        n = 0
        rank1 = ['A']
        for alb in self.wpwa.iter_rows(min_row=1, max_row=self.wpwa.max_row, min_col=1, max_col=self.palastcol, values_only=False):
            n += 1
            if (n % 100) == 0:
                print ('Indexing ', n, ' albums')
            dirName = ''
            pwgId = 0
            if n == 1:
                dirName = ''
                pwgId = 0
            elif n == 1:
                dirName = str(alb[self.padir].value)
                pwgId = int(alb[self.paId].value)
            else:
                rank = str(alb[self.paglobal_rank].value)
                rank2 = rank.split()
                for i in range(len(rank1)):
                    if rank1[i] != rank2[i]:
                        dirName = str(alb[self.padir].value)
                        pwgId = int(alb[self.paId].value)
                rank1 = rank2
                self.setPiwigoAlbRef(dirName, pwgId)

    def setPiwigoAlbRef(self, path, pid):
        
        pwpath = self.padir+1
        pwgid = self.paId+1
        n = 0
        p_id = ''
        for album in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=False):
            n += 1
            if str(album[self.caItem].value) == path:
                if album[self.caPiwigoId].value is not None:
                    print ('ERROR: Double entry on line ', n, ' Piwigo ID col N = ', album[self.caPiwigoId].value, ' for item: ', path)
                else:
                    self.wa.cell(row=n, column=self.caPiwigoId+1).value = pid
                    
    def defineBottomAlbumImages(self):
        n = 1
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            if (n % 100) == 0:
                print ('Checking ', n, ' pics for album images')
            if n > 1 and pic[self.cpAlbumFile].value is not None and pic[self.cpPiwigoId].value is not None and pic[self.cpFileName].value is not None and pic[self.cpDescription].value is not None:
                self.wp.cell(row=n, column=self.cpAlbImgIdId+1).value = pic[self.cpPiwigoId].Value
            n += 1
            
        m = 1
        for pic1 in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            if pic1[self.cpDescription].Value is None pic1[self.cpAlbImgIdId].Value is not None:
                if (m % 100) == 0:
                    print ('Checking ', m, ' albums for album images')
                m += 1
                n = 1
                for pic2 in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
                    if pic1[self.cpParentDoc].value == pic2[self.cpItem].value and pic2[self.cpAlbumFile].value is not None:
                        self.wp.cell(row=n, column=self.cpAlbImgIdId+1).value = pic1[self.cpPiwigoId].Value
                    n += 1

if __name__ == "__main__":
    a = getId(sys.argv)
