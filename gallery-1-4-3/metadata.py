import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
from os import listdir

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Build SQL script for reference sequence numbers in album and files from Gallery

from cols import colsHighStage

class metadataPic(colsHighStage):

    def __init__(self):
        
        super().__init__()
        
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        if self.testMode:
            with open(self.injectdir + self.usersFileName, 'w') as usersFile:
                usersFile.write('')
            with open(self.injectdir + self.userSqlFileName, 'w') as userSqlFile:
                userSqlFile.write('')
            with open(self.injectdir + self.sqlFileName, 'w') as sqlFile:
                sqlFile.write('')
                
        self.metaPics()
        self.metaAlbums()
        
        self.clearUserList()
        
    def metaPics(self):
        n = 0
        sql = ''
        userSql = ''
        users = ''
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n+=1
            picLine = str(self.wp.cell(row=n, column=self.cp+1).value)
            picSourceId = str(self.wp.cell(row=n, column=self.cpItem+1).value)
            picDescr = str(self.wp.cell(row=n, column=self.cpDescription+1).value)
            picWorkspace = str(self.wp.cell(row=n, column=self.cpWorkspace+1).value)
            picEventTime = str(self.wp.cell(row=n, column=self.cpEventTime+1).value)
            picEditBy = str(self.wp.cell(row=n, column=self.cpEditBy+1).value)
            picComment = str(self.wp.cell(row=n, column=self.cpNote+1).value)
            picAlias = str(self.wp.cell(row=n, column=self.cpAlias+1).value)
            picHistory = str(self.wp.cell(row=n, column=self.cpNote2+1).value)
            picFirstTimeStorage = str(self.wp.cell(row=n, column=self.cpDate2+1).value)
            picDate3 = str(self.wp.cell(row=n, column=self.cpDate3+1).value)
            picExif = str(self.wp.cell(row=n, column=self.cpExif+1).value)
            picInitdate = str(self.wp.cell(row=n, column=self.cpInitdate+1).value)
            picParentDoc = str(self.wp.cell(row=n, column=self.cpParentDoc+1).value)
            picFileName = str(self.wp.cell(row=n, column=self.cpFileName+1).value)
            picBareItem = str(self.wp.cell(row=n, column=self.cpBareItem+1).value)
            picAlbumFile = str(self.wp.cell(row=n, column=self.cpAlbumFile+1).value)
            picSeq = self.wp.cell(row=n, column=self.cpSeq+1).value
            picFileError = str(self.wp.cell(row=n, column=self.cpFileError+1).value)
            picDest = str(self.wp.cell(row=n, column=self.cpDest+1).value)
            picPiwigoId = self.wp.cell(row=n, column=self.cpPiwigoId+1).value
            
            if picPiwigoId is not None:
                pw = str(picPiwigoId)
                if pw != '' and n>1:
                    sql = sql + 'update images set name = \'' + picDescr + '\', date_available = \'' + self.toPiwigoDate(picEventTime) + '\', date_metadata_update = now() where id = \'' + pw + '\';\n'
                    if picSeq is not None:
                        sql = sql + 'update image_category set rank = \'' + str(picSeq) + '\' where image_id = \'' + pw + '\';\n'
                    userSql = userSql + 'update images set author = \'' + picEditBy + '\' where id = \'' + pw + '\';\n'
                    users = users + picEditBy + '\n'
                    
            if (n % 100 == 0):
                print (n, ' pictures done')
                
        print 
        with open(self.injectdir + self.usersFileName, 'a') as usersFile:
            usersFile.write(users)
        with open(self.injectdir + self.userSqlFileName, 'a') as userSqlFile:
            userSqlFile.write(userSql)
        with open(self.injectdir + self.sqlFileName, 'a') as sqlFile:
            sqlFile.write(sql)
        
    def metaAlbums(self):
        n = 0
        sql = ''
        userSql = ''
        users = ''
        for alb in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=False):
            n+=1
            albName = str(alb[self.caAlbumName].value)
            albCaItem = str(alb[self.caItem].value)
            albEditBy = str(alb[self.caEditBy].value)
            albParentDoc = str(alb[self.caParentDoc].value)
            albDescription = str(alb[self.caDescription].value)
            albEventTime = str(alb[self.caEventTime].value)
            albInitdate = str(alb[self.caInitdate].value)
            albCa = str(alb[self.ca].value)
            albFileName = str(alb[self.caFileName].value)
            albBareItem = str(alb[self.caBareItem].value)
            albSeq = str(alb[self.caSeq].value)
            albMD5 = str(alb[self.caMD5].value)
            albPiwigoId = str(alb[self.caPiwigoId].value)
            albAlbumImg = str(alb[self.caAlbumImg].value)
            albAlbumPath = str(alb[self.caAlbumPath].value)

            if albPiwigoId is not None:
                pw = str(albPiwigoId)
                if pw != '' and n>1:
                    sql = sql + 'update categories set name = \'' + albDescription + '\''
                    if albSeq is not None:
                        sql += ', rank = \'' + albSeq + '\''
                    sql += ' where id = \'' + pw +  '\';\n'
                    if albSeq is not None:
                        sql = sql + 'update image_category set rank = \'' + str(picSeq) + '\' where image_id = \'' + pw + '\';\n'
                    #userSql = userSql + 'update images set author = \'' + picEditBy + '\' where id = \'' + pw + '\';\n'
                    #users = users + picEditBy + '\n'
                    
            if (n % 100 == 0):
                print (n, ' albums done')
                
        with open(self.injectdir + self.usersFileName, 'a') as usersFile:
            usersFile.write(users)
        with open(self.injectdir + self.userSqlFileName, 'a') as userSqlFile:
            userSqlFile.write(userSql)
        with open(self.injectdir + self.sqlFileName, 'a') as sqlFile:
            sqlFile.write(sql)
    
    def metaPics(self):
        n = 0
        sql = ''
        userSql = ''
        users = ''
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic , values_only=False):
            n+=1
			picDescription = str(pic[self.cpDescription].value)
			picFileType = str(pic[self.cpFileType].value)
			picItem = str(pic[self.cpItem].value)
			picFileName = str(pic[self.cpFileName].value)
			picParentDoc = str(pic[self.cpParentDoc].value)
			picEditBy = str(pic[self.cpEditBy].value)
			picComment = str(pic[self.cpComment].value)
			picAlbumFile = str(pic[self.cpAlbumFile].value)
			picInitdate = str(pic[self.cpInitdate].value)
			picKeyWord = str(pic[self.cpKeyWord].value)
			picSeq = str(pic[self.cpSeq].value)
			picCp = str(pic[self.cp].value)
			picBareItem = str(pic[self.cpBareItem].value)
			picFileError = str(pic[self.cpFileError].value)
			picDest = str(pic[self.cpDest].value)
			picPiwigoId = str(pic[self.cpPiwigoId].value)
			picMigrInfo = str(pic[self.cpMigrInfo].value)
			picPath = str(pic[self.cpPath].value)

            if pic[self.cpPiwigoId].value is not None:
                if picPiwigoId != '' and n>1:
                    sql = sql + 'update images set name = \'' + picDescription + '\''
                    if pic[self.cpSeq].value is not None:
                        sql += ', rank = \'' + picSeq + '\''
                    sql += ' where id = \'' + picPiwigoId +  '\';\n'
                    if pic[self.cpSeq].value is not None:
                        sql = sql + 'update image_category set rank = \'' + picSeq + '\' where image_id = \'' + picPiwigoId + '\';\n'
                    #userSql = userSql + 'update images set author = \'' + picEditBy + '\' where id = \'' + picPiwigoId + '\';\n'
                    #users = users + picEditBy + '\n'
                    
            if (n % 100 == 0):
                print (n, ' pics done')
                
        with open(self.injectdir + self.usersFileName, 'a') as usersFile:
            usersFile.write(users)
        with open(self.injectdir + self.userSqlFileName, 'a') as userSqlFile:
            userSqlFile.write(userSql)
        with open(self.injectdir + self.sqlFileName, 'a') as sqlFile:
            sqlFile.write(sql)

    
    def toPiwigoDate (self, d):
        
        dl = d.split('-')
        if len(dl) < 3:
            print ('ERROR, feil lengde: ', d)
        r = dl[0] + '.' + dl[1] + '.' + dl[2] + ' 12:00'
        return r
    
    def clearUserList(self):
        b = 3
        lines_seen = set()
        outfile = open(self.injectdir + 'unique_' + self.usersFileName, "w")
        for line in open(self.injectdir + self.usersFileName, "r"):
            if line not in lines_seen:
                outfile.write(line)
                lines_seen.add(line)
        outfile.close()
        
        if os.path.exists(self.injectdir + self.usersFileName):
            os.remove(self.injectdir + self.usersFileName)
    
    
a = metadataPic()
