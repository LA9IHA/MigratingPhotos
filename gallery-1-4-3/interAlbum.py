import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
from os import listdir

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Set up references between albums in Gallery
# it's about assigning parent piwigo ID into PIC

from cols import cols

class albumRefs(cols):

    def __init__(self):
        
        super().__init__()
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        runPic = False
        runAlb = False
        if len(pa) >= 2:
            if pa[1] == 'P':
                runPic = True
            elif pa[1] == 'A':
                runAlb = True
        else:
            runPic = True
            runAlb = True
        
        if runAlb:
            self.linkAlbums()
            
        if runPic:
            self.linkPics()
        
        if runAlb:
            self.album_wb.save(self.subdir + 'interAlbum1.xlsx')
        if runPic:
            self.pic_wb.save(self.subdir + self.fOutputPic)

    def linkAlbums(self):
        
        n = 0
        for alb in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=False):
            if n > 0 and alb[self.caParentPwgId].value is None:
                if (n % 100) == 0:
                    print ('Linked ', n, ' albums')
                i = 0
                for alb2 in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=False):
                    if i > 1 and alb2[self.caPiwigoId].value is not None:
                        if alb[self.caParentDoc].value == alb2[self.caItem].value:
                            self.wa.cell(row=n+1, column=self.caParentPwgId+1).value = alb2[self.caPiwigoId].value
                            print (n, ' Child: ',  alb[self.caPiwigoId].value, ' parent: ', alb2[self.caPiwigoId].value)
                    i+=1
            n+=1
                


    def linkPics(self):
        n = 0
        sql = ''
        userSql = ''
        users = ''
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            if n > 0 and pic[self.cpPiwigoId].value is not None:
                if pic[self.cpParentDoc].value is not None:
                    m = 0
                    for alb in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
                        if m > 0 and alb[self.caPiwigoId].value is not None:
                            if pic[self.cpParentDoc].value == alb[self.caItem].value:
                                self.wp.cell(row=n, column=self.cpParentPwgId+1).value = int(alb[self.caPiwigoId].value)
                                print (n, ' Child: ', pic[self.cpPiwigoId].value, ' --- Parent: ', alb[self.caPiwigoId].value)
                        m+=1

            n+=1
            if (n % 100 == 0):
                print (n, ' pictures done')
                  
if __name__ == "__main__":
    a = albumRefs(sys.argv)

