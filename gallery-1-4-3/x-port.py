import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib
import re
from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
#from datetime import datetime, timedelta

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Export from Gallery album to Piwigo Fils structure. Build metadata from items


from cols import cols

class xport(cols):

    def __init__(self):
        
        super().__init__()

        #self.path_depth = 0
        
        if os.path.exists(self.treedir):
            shutil.rmtree(self.treedir)
        os.makedirs(self.treedir)
        
        for filname in os.listdir(self.treedir):
            filepath = os.path.join(self.treedir, filname)
            try:
                if os.path.isfile(filepath) or os.path.islink(filepath):
                    os.unlink(filepath)
                elif os.path.isdir(filepath):
                    shutil.rmtree(filepath)
            except Exception as e:
                print(f'Could not purge {filepath}')
                
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        self.lineNo = 0

        if not os.path.exists(self.errLog):
            with open(self.errLog, 'w') as logfile:
                logfile.write('x-port errs')
            logfile.close()
        
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]

        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        print ('x-port, Gallery exporting from: ', self.topParent)

        self.makeTopAlbums(self.wa)
        
        self.album_wb.save(self.subdir + self.fOutputAlbum)
        self.pic_wb.save(self.subdir + self.fOutputPic)

    def makeTopAlbums (self, sheet):
            
        for album in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if album[self.caParentDoc] == self.topParent:
                self.createTree(album)

    # Create a directory with sub-directories and photos
    def createTree(self, row):
        #self.path_depth += 1
        self.createDir(row)
        self.createPics(row)
        self.createChildren(row)
        #self.path_depth -= 1
        
    # Create child directories under a parent dirctory
    def createChildren(self, r):
        for row in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if row[self.caParentDoc] == r[self.caItem]:
                self.createTree(row)
        
    def createDir(self, r):
        
        thisPath = self.treedir + r[self.caAlbumPath]
        if not os.path.exists(thisPath):
            os.makedirs(thisPath)

    def createPics(self, r):
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=True):
            if pic[self.cpItem] is not None and r[self.caItem] is not None and pic[self.cpFileName] is not None and pic[self.cpPath] is not None:
                if r[self.caItem] == pic[self.cpParentDoc]:
                    sourcePath = self.subdir + 'PHOTOS' + pic[self.cpFileName]
                    destPath = self.treedir + pic[self.cpPath]
                    self.lineNo += 1
                    #print ('FROM: ', sourcePath, '  ---- TO: ', destPath)
                    if (self.lineNo % 100 == 0):
                        print ('Copied ', self.lineNo, ' files to export directory ', self.treedir)
                    if '.' + pic[self.cpPath].split(".")[-1] in self.custom_image_extensions:
                        try:
                            shutil.copy(sourcePath, destPath)
                        except Exception as e:
                            print ('FAILED COPY ', sourcePath, ' TO ', destPath, ' ERROR MSG: ', e)
                            with open(self.errLog) as logfile:
                                errmsg = 'FAILED COPY ', sourcePath, ' TO ', destPath, ' >>> ERROR MSG: ', e, '\n'
                                logfile.write(errmsg)
                    else:
                        print ('.' + pic[self.cpPath].split(".")[-1], '  #### ', 'Cannot copy ', sourcePath, '--- Not in ', self.custom_image_extensions)
    
a = xport()