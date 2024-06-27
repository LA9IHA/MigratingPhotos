import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
from os import listdir

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Set up references between albums in Gallery
# it's about assigning parent piwigo ID into PIC

from cols import cols

class albumRefs(cols):

    def __init__(self):
        
        super().__init__()
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        interAlbumfile = self.subdir + self.fInterAlbum
        self.album_iwa = load_workbook(filename=interAlbumfile)
        self.iwa = self.album_iwa.worksheets[0]
        
        self.linkAlbums()
        
        self.pic_wb.save(self.subdir + self.fOutputPic)
        
    def linkAlbums(self):
        
        n = 1
        hits = 0
        nextLocation = 1
        newItems = 0
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            if n > 1 and pic[self.cpDescription].value is None:
                for ref in self.iwa.iter_rows(min_row=1, max_row=self.iwa.max_row, min_col=1, max_col=self.iaLastCol, values_only=False):
                    if pic[self.cpParentDoc].value == ref[self.iaParentDoc].value and ref[self.iaItem].value is not None:
                        hits += 1
                        if hits == nextLocation:
                            self.wp.cell(row=n, column=self.cpItem+1).value = str(ref[self.iaItem].value)
                            nextLocation += 1
                            newItems += 1
                            #print (n, ' Child: ', pic[self.cpItem].value, ' --- Parent: ', pic[self.cpParentDoc].value)
                    else:
                        hits = 0
                        nextLocation = 1
            n+=1
            if (n % 100 == 0):
                print (n, ' pics checked')
        print ('\nDone, ', newItems, ' items added to', self.fOutputPic)
                  
a = albumRefs()