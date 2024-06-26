import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
from os import listdir

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Set up references between albums in Gallery
# it's about assigning parent piwigo ID into PIC

from cols import colsHighStage

class albumRefs(colsHighStage):

    def __init__(self):
        
        super().__init__()
        
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
                        
        self.linkPics()
        
        self.pic_wb.save(self.subdir + self.fOutputPic)
        
    def linkPics(self):
        n = 0
        sql = ''
        userSql = ''
        users = ''
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n+=1
            
            if pic[self.cpPiwigoId].value is not None:
                if pic[self.cpPiwigoId].value != '' and n>1:
                    for alb in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastPic, values_only=False):
                        if pic[self.cpParentDoc].value is not None and alb[self.caPiwigoId].value is None and pic[self.cpParentDoc].value == alb[self.caItem].value:
                             self.wp.cell(row=n, column=self.cpPiwigoParentId+1).value = int(alb[self.caPiwigoId].value)
                    
            if (n % 100 == 0):
                print (n, ' pictures done')
                  
    
a = albumRefs()
