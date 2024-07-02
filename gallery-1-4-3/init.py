import openpyxl
from openpyxl import Workbook
#from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
import sys
import os
import shutil
import datetime as dt
from datetime import datetime, timedelta

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Initiate metadata files Album.xlsx and Pic.xlsx


from cols import cols

class init(cols):

    def __init__(self, phaseNo):
        
        super().__init__()

        self.path_depth = 0
        self.lines = 1
        
        if os.path.exists(self.treedir):
            shutil.rmtree(self.treedir)
        os.makedirs(self.treedir)
        
        for filname in os.listdir(self.treedir):
            filepath = os.path.join(self.treedir, filname)
            try:
                if os.path.isfile(filepath) or os.path.islink(filepath):
                    os.unlink(filepath)
                elif os.path.isdir(filepath):
                    shutil.rmtree(filepath)
            except Exception as e:
                print(f'Could not purge {filepath}')
                
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
            
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        phase = int(phaseNo)
        print ('Phase: ', phase)
        if phase == 1:
            print ('     Perform some housekeeping on Album.xlsx and Pic.xlsx')
            self.addSeqPic()
            self.addSeqAlbum()
        elif phase == 2:
            print ('     Define source location of directories, add line numbers on Albums')
            self.albums2path('0', '')
        elif phase == 3:
            print ('     Define source location of files, add line numbers on Photos')
            self.pic2path()
        elif phase == 4:
            print ('     Check and fix paths in Pic.xlsx')
            self.fixPath()
        else:
            print ('Unknown phase: ', phaseNo)
            
        if phase in [1, 2]:
	        self.album_wb.save(self.subdir + self.fOutputAlbum)
        if phase in [1, 3, 4]:
	        self.pic_wb.save(self.subdir + self.fOutputPic)
        
    def addSeqPic (self):
        n = 1
        for row in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=True):
            if n > 1:
                self.wp.cell(row=n, column=self.cp+1).value = n-1
            if (n % 100) == 0:
                print('Indexed ', n, ' photos')
            n += 1
        c = 1
        for col in self.colsPic:
            self.wp.cell(row=1, column=c).value = col
            c+=1

        print ('  Done, saving at ', self.subdir + self.fOutputPic)
        
    def addSeqAlbum (self):
        
        n = 1
        for album in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if n > 1:
                self.wa.cell(row=n, column=self.ca+1).value = n-1
                
                start_date = dt.datetime.strptime('1970-01-01 00:00:00', '%Y-%m-%d %H:%M:%S')
                seconds_since_epoch = album[self.caEventTime]
                event_time = start_date + dt.timedelta(seconds=seconds_since_epoch)
                self.wa.cell(row=n, column=self.caEventTime+1).value = event_time
            if (n % 100) == 0:
                print('Indexed ', n, ' albums')
            n += 1
        a = 1
        for col in self.colsAlbum:
            self.wa.cell(row=1, column=a).value = col
            a = a + 1
            
        print ('  Done, saving at ', self.subdir + self.fOutputAlbum)

    def albums2path (self, parent, path):
        
        for album in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if str(album[self.caParentDoc]) == parent:
                if (self.lines % 100) == 0:
                    print ('Added path to ', self.lines, ' albums')
                self.lines += 1
                thisPath = path + album[self.caItem] + '/'
                self.wa.cell(row=album[self.ca]+1, column=self.caAlbumPath+1).value = thisPath
                self.albums2path(album[self.caItem], thisPath)
    
    def pic2path (self):
        
        self.lines = 1
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=True):
            for album in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
                if str(pic[self.cpParentDoc]) == str(album[self.caItem]):
                    if pic[self.cpItem] is not None:
                        self.wp.cell(row=int(pic[self.cp]+1), column=self.cpPath+1).value = album[self.caAlbumPath] + pic[self.cpItem] + '.' + pic[self.cpFileType]
            if (self.lines % 100) == 0:
                print ('Added path to ', self.lines, ' photos')
            self.lines += 1
           
    def fixPath (self):
        
        self.lines = 1
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=True):
            if pic[self.cpDescription] is not None and pic[self.cpPath] is None:
            	filnameDirs = pic[self.cpFileName].value.split("/")
            	if filnameDirs[1] != pic[self.cpParentDoc] and len(filnameDirs[2]) > 2:
            		self.wp.cell(row=self.lines), column=self.cpFileName+1).value = '/' + pic[self.cpParentDoc] + '/' + filnameDirs[2]
            if (self.lines % 100) == 0:
                print ('Added path to ', self.lines, ' photos')
            self.lines += 1
           
           
if __name__ == "__main__":
    if len(sys.argv) > 1:
        a = init(sys.argv[1])
    else:
        print ('\n\nMissing phase no. Add phase no as follows:\n')
        print ('python3 init.py 1 : Run phase 1 to perform some housekeeping on Album.xlsx and Pic.xlsx')
        print ('python3 init.py 2 : Run phase 2 to define source location of directories, add line numbers on Albums')
        print ('python3 init.py 3 : Run phase 3 to define source location of files, add line numbers on Photos\n')
        print ('python3 init.py 4 : Run phase 4 to check and fix paths in Pic.xlsx\n')
