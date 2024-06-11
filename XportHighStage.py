import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib
import re

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook

# Purpose: Prepare HighStage album with photos for  igration to Piwigo
# Pre requisites: Album.xlsx and Pic.xlsx is created from Highstage
# Licence: GNU 2.0
# Author: Ottar Kvindesland, 2024
# Reference: https://piwigo.miraheze.org/wiki/HighstageExport

class XportHighStage:

    def __init__(self, h, t, s):
        self.homedir = h
        self.treedir = t
        self.subdir = s
        self.path_depth = 0
        
        # Define top parent in HighStage. If blanks, update Album.xlsx and replace empty
        # parent album names with ZZZ and name topParent ZZZ.
        
        #self.topParent = 'SPC1064-1A'
        self.topParent = 'ALBUM1732-1A'
                
        if os.path.exists(self.treedir):
            shutil.rmtree(self.treedir)
        os.makedirs(self.treedir)
        
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        self.path = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''] 
        self.path[0] = self.treedir
        self.MDerr = 'UFFDA'
        
        self.custom_image_extensions = ['.jpeg', '.png', '.jpg', '.gif']
        self.custom_substitutions = [('&', 'et'), ('(',''), (')',''), (' !', ''), ("'", ' '), (' ', '_'), (',', '_')]
        self.custom_video_mime_types = ['media', 'video']
        
        picfile = self.subdir + 'Pic.xlsx'
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + 'Album.xlsx'
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]

        # Columns used in Album.xlsx
        self.ca = 0 # A = Id
        self.caItem = 1 # B = Item
        self.caDescription = 2 # C = Description
        self.caWorkspace = 3 # D = Workspace, i.e. access group
        self.caEventTime = 4 # etc....
        self.caEditBy = 5
        self.caNote = 6
        self.caInitdate = 7
        self.caParentDoc = 8
        self.caFileName = 9
        self.caBareItem = 10
        self.caSeq = 11
        self.caMD5 = 12
        self.caLastAlbum = 13
        
        # Columns used in Pic.xlsx
        self.cp = 0 # A = ID
        self.cpItem = 1 # B = Item
        self.cpDescription = 2 # etc ...
        self.cpWorkspace = 3
        self.cpEventTime = 4
        self.cpEditBy = 5
        self.cpNote = 6
        self.cpAlias = 7
        self.cpNote2 = 8 # History
        self.cpDate2 = 9  # First time storage
        self.cpDate3 = 10
        self.cpExif = 11
        self.cpInitdate = 12 # Date taken
        self.cpParentDoc = 13
        self.cpFileName = 14
        self.cpBareItem = 15
        self.cpAlbumFile = 16
        self.cpSeq = 17  # Q Sequence
        self.cpFileError = 18 # R - Error message
        self.cpLastPic = 19
        
        self.makeTopAlbums(self.wa)
        self.album_wb.save(self.subdir + 'Album1.xlsx')
        self.pic_wb.save(self.subdir + 'Pic1.xlsx')
        
    def makeTopAlbums (self, sheet):
        for row in sheet.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if row[self.caParentDoc] == self.topParent:
                self.createTree(row)
                
    # Create a directory with sub-directories and photos
    def createTree(self, row):
        self.path_depth += 1
        self.path[self.path_depth] = row[self.caDescription]
        md = self.createAMD5(row)
        self.createDir(row)
        self.createPics(row, md)
        self.createChildren(row, md)
        self.path_depth -= 1
        
    # Define an MD5 hash for album to recognize album pictures
    def createAMD5(self, r):
        md = self.MDerr
        if (r[self.caFileName] != ''):
            rnum = int(r[self.ca]) + 1
            fdir = self.subdir + 'ALBUMS/' + r[self.caBareItem] + '/' + r[self.caBareItem] + '/' + r[self.caFileName]
            md = hashlib.md5(open(fdir,'rb').read()).hexdigest()
            self.wa.cell(row=rnum, column=self.caMD5).value = md
        return md
        
    # Create child directories under a parent dirctory
    def createChildren(self, r, mdParent):
        for row in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if row[self.caParentDoc] == r[self.caItem]:
                albFile = ''
                md = self.createAMD5(row)
                if md == mdParent:
                    albFile = 'Y'
                print ('CHILD ALBUM: ', row[self.caItem], ' - ', albFile)
                self.createTree(row)
       
        
    def createDir(self, row):
        subtreepath = ''
        for n in range (self.path_depth + 1):
            subtreepath = subtreepath + self.path[n]
            if n>0:
                subtreepath = subtreepath + '/'
        os.makedirs(self.cleanPathName(subtreepath))

    def createPics(self, r, m):
        n = 0
        children = 0
        c = self.cpParentDoc+1
        bi = self.cpBareItem+1
        fn = self.cpFileName+1
        # ds = self.caDescription+1
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n+=1
            parent = self.wp.cell(row=n, column=c).value
            bareItem = self.wp.cell(row=n, column=bi).value
            fileName = self.wp.cell(row=n, column=fn).value
            if parent == r[self.caItem]:
                children += 1
                fpath = self.subdir + 'PHOTOS/' + bareItem + '/' + bareItem + '/'
                fthumb = fpath + 'doc_pic.jpg'
                ffull = fpath + fileName
                md = hashlib.md5(open(fthumb,'rb').read()).hexdigest()
                if md == m:
                    self.wp.cell(row=n, column=self.cpAlbumFile).value = 'AlbumImage'
                self.copyFiles(r, ffull, fileName)
        print(children, 'bilder i', r[self.caItem], '-', r[self.caDescription])

    def copyFiles(self, row, origFile, fileName):
        subtreepath = ''
        for n in range (self.path_depth + 1):
            subtreepath = subtreepath + self.path[n]
            if n>0:
                subtreepath = subtreepath + '/'
        dest = self.cleanPathName(subtreepath + fileName)
        shutil.copy(origFile, dest)

    def norskeBokstaver(self, name):
        
        name = name.replace("&#230;", "æ")
        name = name.replace("&#248;", "ø")
        name = name.replace('&#229;', "å")
        name = name.replace("&#198;", "Æ")
        name = name.replace("&#216;", "Ø")
        name = name.replace("&#197;", "Å")
        
        return name

    def cleanPathName(self, name):
        """Substitue some special characters"""
        name = self.norskeBokstaver(name)
        for a,b in self.custom_substitutions:
            name = name.replace(a,b)
        name = re.sub(r'[^a-zA-Z0-9]', '_', name)
        
        #name = name.decode('iso-8859-1')
        return name


homedir = "/Volumes/home/Oppgaver/hstransfer/" 
treedir = homedir + "tree/"
subdir = homedir + "source/"
    
a = XportHighStage(homedir, treedir, subdir)