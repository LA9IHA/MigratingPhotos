# See ../highstage/cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Export from HighStage album to Piwigo Fils structure. Build metadata from items



import datetime
import os
import shutil
import hashlib
import pandas as pd
import fnmatch
import glob
import zipfile

import mysql.connector

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
from os import listdir

# Purpose: Check the integrity of the media files
# Licence: GNU 2.0
# Copyright: Ottar Kvindesland, 2024

class extractTables ():

    def __init__(self):
        super().__init__()
        
        self.dbName = 'album'
        self.host='localhost'
        self.user='migrationuser'
        self.password='secretPassword'
        self.migrateDir = '/tmp/'
            
        tables = self.getAllTableNames()
        zipFiles = []
        print ('Extracting tables\nWhen complete run the script lines below:\n\n')
        for table in tables:
            self.dumpTable (table[0])
            self.genXlsx (table[0])
            zipFiles.add(table[0] + '.xlsx')
            
        with zipfile.ZipFile('dbdump.zip', 'w') as dumpzip:
            for fileName in zipFiles:
                dumpzip.write(fileName)
        
    def getAllTableNames(self):
        
        conn = mysql.connector.connect(
            host=self.host,
            user=self.user,
            password=self.password,
            database=self.dbName
            )
        cursor = conn.cursor()
        cursor.execute('show tables')
        results = cursor.fetchall()
        r = []
        for row in results:
            r.append(row)
            
        return r

    def dumpTable (self, table):
        
        c1 = "SET @col_names = (SELECT GROUP_CONCAT(CONCAT('\\'', COLUMN_NAME, '\\'')) FROM INFORMATION_SCHEMA.COLUMNS WHERE TABLE_NAME = '" + table + "' AND TABLE_SCHEMA = '" + self.dbName + "');"
        c2 = "SET @sql = CONCAT('SELECT ', @col_names, ' UNION ALL SELECT * FROM " + table + " INTO OUTFILE \\'" +  self.migrateDir + table + ".csv\\' FIELDS TERMINATED BY \\\',\\\' ENCLOSED BY \\\'\"\\\' LINES TERMINATED BY \\\'\n\\\'');"
        c3 = 'PREPARE stmt FROM @sql;'
        c4 = 'EXECUTE stmt;'
        c5 = 'DEALLOCATE PREPARE stmt;'

        conn = mysql.connector.connect(
            host=self.host,
            user=self.user,
            password=self.password,
            database=self.dbName
            )
        cursor = conn.cursor()
        cursor.execute(c1)
        cursor.execute(c2)
        cursor.execute(c3)
        cursor.execute(c4)
        cursor.execute(c5)
        
    def getCSVfiles(self):
        
        results = [file for file in os.listdir(self.migrateDir) if fnmatch.fnmatch(file, '*.csv')]
        r = []
        for row in results:
            r.append(row)
            
        return r

    def genXlsx (self, tableName):
        
        table = tableName[0]
        csvFileName = self.migrateDir + tableName + '.csv'
        wb = Workbook()
        ws = wb.active
        r = 1
        print ('sudo rm ' + csvFileName )
        with open(csvFileName, 'r') as csvFile:
            for line in csvFile:
                line = line.strip().replace(',\\N', ',""')
                elements = line.split('","')
                c = 1
                for element in elements:
                    element = element.strip('"')
                    ws.cell(row=r, column=c).value = element
                    c += 1
                r += 1
        wb.save(tableName + '.xlsx')
  
a = extractTables()


