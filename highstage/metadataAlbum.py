import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
from os import listdir

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Build reference sequence numbers in album and files

from cols import colsHighStage

class metadataPic(colsHighStage):

    def __init__(self):
        
        super().__init__()
        
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        if self.testMode:
            with open(self.subdir + self.usersFileName, 'w') as usersFile:
                usersFile.write('')
            with open(self.subdir + self.userSqlFileName, 'w') as userSqlFile:
                userSqlFile.write('')
            with open(self.subdir + self.sqlFileName, 'w') as sqlFile:
                sqlFile.write('')
                
        #self.metaPics()
        self.metaAlbums()
        
        self.clearUserList()
        
    def metaAlbums(self):
        n = 0
        sql = ''
        userSql = ''
        users = ''
        for pic in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=False):
            n+=1
            albCa = str(self.wa.cell(row=n, column=self.ca+1).value)
            albCaItem = str(self.wa.cell(row=n, column=self.caItem+1).value)
            albDescription = str(self.wa.cell(row=n, column=self.caDescription+1).value)
            albWorkspace = str(self.wa.cell(row=n, column=self.caWorkspace+1).value)
            albEventTime = str(self.wa.cell(row=n, column=self.caEventTime+1).value)
            albEditBy = str(self.wa.cell(row=n, column=self.caEditBy+1).value)
            albNote = str(self.wa.cell(row=n, column=self.caNote+1).value)
            albInitdate = str(self.wa.cell(row=n, column=self.caInitdate+1).value)
            albParentDoc = str(self.wa.cell(row=n, column=self.caParentDoc+1).value)
            albFileName = str(self.wa.cell(row=n, column=self.caFileName+1).value)
            albBareItem = str(self.wa.cell(row=n, column=self.caBareItem+1).value)
            albSeq = self.wa.cell(row=n, column=self.caSeq+1).value
            albMD5 = str(self.wa.cell(row=n, column=self.caMD5+1).value)
            albPiwigoId = self.wa.cell(row=n, column=self.caPiwigoId+1).value
            albAlbumImg = str(self.wa.cell(row=n, column=self.caAlbumImg+1).value)
            albAlbumPath = str(self.wa.cell(row=n, column=self.caAlbumPath+1).value)

            if albPiwigoId is not None:
                pw = str(albPiwigoId)
                if pw != '' and n>1:
                    sql = sql + 'update categories set name = \'' + albDescription + '\''
                    if albSeq is not None:
                        sql += ', rank = \'' + albSeq + '\''
                    sql += ' where id = \'' + pw +  '\';\n'
                    if albSeq is not None:
                        sql = sql + 'update image_category set rank = \'' + str(picSeq) + '\' where image_id = \'' + pw + '\';\n'
                    #userSql = userSql + 'update images set author = \'' + picEditBy + '\' where id = \'' + pw + '\';\n'
                    #users = users + picEditBy + '\n'
                    
            if (n % 100 == 0):
                print (n, ' pictures done')
                
        with open(self.subdir + self.usersFileName, 'a') as usersFile:
            usersFile.write(users)
        with open(self.subdir + self.userSqlFileName, 'a') as userSqlFile:
            userSqlFile.write(userSql)
        with open(self.subdir + self.sqlFileName, 'a') as sqlFile:
            sqlFile.write(sql)
            
    def toPiwigoDate (self, d):
        
        dl = d.split('-')
        if len(dl) < 3:
            print ('ERROR, feil lengde: ', d)
        r = dl[0] + '.' + dl[1] + '.' + dl[2] + ' 12:00'
        return r
    
    def clearUserList(self):
        b = 3
        lines_seen = set()
        outfile = open(self.subdir + 'unique_' + self.usersFileName, "w")
        for line in open(self.subdir + self.usersFileName, "r"):
            if line not in lines_seen:
                outfile.write(line)
                lines_seen.add(line)
        outfile.close()
        
        if os.path.exists(self.subdir + self.usersFileName):
            os.remove(self.subdir + self.usersFileName)
    
a = metadataPic()
