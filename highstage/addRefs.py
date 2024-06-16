import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
from os import listdir

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Build reference sequence numbers in album and files

from cols import colsHighStage

class addSeqs(colsHighStage):

    def __init__(self):
        
        super().__init__()
        
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        referencesfile = self.subdir + self.fRefList
        self.references_wb = load_workbook(filename=referencesfile)
        self.wr = self.references_wb.worksheets[0]
        
        self.lineCounter = 0
        
        self.refPics()
        self.pic_wb.save(self.subdir + self.fOutputPic)
        self.refAlbums()
        self.getAlbumImages(self.topParent)
        self.album_wb.save(self.subdir + 'Album1.xlsx')

        
    def refPics(self):
        n = 0
        c = self.cpParentDoc+1
        bi = self.cpBareItem+1
        fn = self.cpFileName+1
        err = self.cpFileError+1
        sec = 0
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n+=1
            parent = self.wp.cell(row=n, column=c).value
            bareItem = self.wp.cell(row=n, column=bi).value
            
            if parent is not None and bareItem is not None:
                seq = self.getSeq(parent, bareItem)
                if (seq > 0):
                    self.wp.cell(row=n, column=self.cpSeq+1).value = seq
                
            if (n % 100 == 0):
                print (n, ' pictures done')
        
        i = 1
        for col in self.colsPic:
            self.wp.cell(row=1, column=i).value = col
            i = i + 1
                    
    def refAlbums(self):
        n = 0
        c = self.caParentDoc+1
        bi = self.caBareItem+1
        sec = 0
        
        for album in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=False):
            n+=1
            parent = self.wa.cell(row=n, column=c).value
            bareItem = self.wa.cell(row=n, column=bi).value
            
            if parent is not None and bareItem is not None:
                seq = self.getSeq(parent, bareItem)
                if (seq > 0):
                    self.wa.cell(row=n, column=self.caSeq+1).value = seq

            if (n % 100 == 0):
                print (n, ' albums done')
                
        i = 1
        for col in self.colsAlbum:
            self.wa.cell(row=1, column=i).value = col
            i = i + 1
            
    def getSeq(self, p, c):
        
        p = p.split('-')[0]
        c = c.split('-')[0]
        
        n = 0
        cp = self.rpParent+1
        cc = self.rpChild+1
        cs = self.rpSeq+1
        hits = 0
        s = 0
        
        for ref in self.wr.iter_rows(min_row=1, max_row=self.wr.max_row, min_col=1, max_col=self.rpLastCol, values_only=False):
            n+=1
            parent = self.wr.cell(row=n, column=cp).value.split('-')[0]
            child = self.wr.cell(row=n, column=cc).value.split('-')[0]
            
            if (p == parent and c == child):
                hits = hits + 1
                s = int(self.wr.cell(row=n, column=cs).value)
        return s

    def getAlbumImages(self, p):
        
        for parent in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if parent[self.caItem] == p:
                for child in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
                    if child[self.caParentDoc] == p:
                        self.lineCounter += 1
                        if (self.lineCounter % 10 == 0):
                            print (self.lineCounter, ' albums images defined')
                        alImg = ''
                        if parent[self.caMD5] == child[self.caMD5]:
                            self.wa.cell(row=int(child[self.ca])+1, column=self.caAlbumImg+1).value = 'AlbumImage'
                        self.getAlbumImages(child[self.caItem])
    
a = addSeqs()
