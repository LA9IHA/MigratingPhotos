import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib
import re

# See cols.py header for info.
# (C) 2024: Ottar Kvindesland, Licence: GPL 2.0
# Purpose: Export from HighStage album to Piwigo Fils structure. Build metadata from items

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook

from cols import colsHighStage

class xport(colsHighStage):

    def __init__(self):
        
        super().__init__()

        self.path_depth = 0
        
        if os.path.exists(self.treedir):
            shutil.rmtree(self.treedir)
        os.makedirs(self.treedir)
        
        for filname in os.listdir(self.treedir):
            filepath = os.path.join(self.treedir, filname)
            try:
                if os.path.isfile(filepath) or os.path.islink(filepath):
                    os.unlink(filepath)
                elif os.path.isdir(filepath):
                    shutil.rmtree(filepath)
            except Exception as e:
                print(f'Could not purge {filepath}')
                
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]

        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        self.makeTopAlbums(self.wa)
        
        self.album_wb.save(self.subdir + self.fOutputAlbum)
        self.pic_wb.save(self.subdir + self.fOutputPic)
        
    def makeTopAlbums (self, sheet):
        c = 1
        for col in self.colsPic:
            self.wp.cell(row=1, column=c).value = col
            c = c + 1
        a = 1
        for col in self.colsAlbum:
            self.wa.cell(row=1, column=a).value = col
            a = a + 1
            
        for row in sheet.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if row[self.caParentDoc] == self.topParent:
                m = self.createTree(row)
                
    # Create a directory with sub-directories and photos
    def createTree(self, row):
        self.path_depth += 1
        self.path[self.path_depth] = self.cleanPathName(row[self.caDescription])
        md = self.createAMD5(row)
        self.createDir(row)
        self.createPics(row, md)
        self.createChildren(row, md)
        self.path_depth -= 1
        
        return md
        
    # Define an MD5 hash for album to recognize album pictures
    def createAMD5(self, r):
        md = self.MDerr
        if (r[self.caFileName] != ''):
            rnum = int(r[self.ca]) + 1
            fdir = self.subdir + 'ALBUMS/' + r[self.caBareItem] + '/' + r[self.caBareItem] + '/' + r[self.caFileName]
            md = hashlib.md5(open(fdir,'rb').read()).hexdigest()
            self.wa.cell(row=rnum, column=self.caMD5+1).value = md
        return md
        
    # Create child directories under a parent dirctory
    def createChildren(self, r, mdParent):
        for row in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if row[self.caParentDoc] == r[self.caItem]:
                albFile = ''
                md = self.createAMD5(row)
                if md == mdParent:
                    albFile = 'Y'
                    #
                    # FIX THIS! Need to mark album files of child album
                    #
                    self.wa.cell(row=rnum, column=self.caAlbumImg+1).value = 'ALBUM FILE'
                print ('CHILD ALBUM: ', row[self.caItem], ' - ', albFile)
                self.createTree(row)
       
        
    def createDir(self, r):
        subtreepath = ''
        for n in range (self.path_depth + 1):
            subtreepath = subtreepath + self.path[n]
            if n>0:
                subtreepath = subtreepath + '/'
        os.makedirs(subtreepath)
        
        if (r[self.caFileName] != ''):
            rnum = int(r[self.ca]) + 1
            if subtreepath.startswith(self.homedir):
                relPath = subtreepath[len(self.homedir):]
            self.wa.cell(row=rnum, column=self.caAlbumPath+1).value = relPath

    def createPics(self, r, m):
        n = 0
        children = 0
        c = self.cpParentDoc+1
        bi = self.cpBareItem+1
        fn = self.cpFileName+1
        dst = self.cpDest+1
        # ds = self.caDescription+1
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n+=1
            dest = ''
            parent = self.wp.cell(row=n, column=c).value
            bareItem = self.wp.cell(row=n, column=bi).value
            fileNameRaw = self.wp.cell(row=n, column=fn).value
            prefix, _, suffix = fileNameRaw.rpartition('.')
            fileName = self.cleanPathName(prefix) + '.' + self.cleanPathName(suffix)
            if parent == r[self.caItem]:
                children += 1
                fpath = self.subdir + 'PHOTOS/' + bareItem + '/' + bareItem + '/'
                fthumb = fpath + 'doc_pic.jpg'
                fileOrig = fpath + fileNameRaw
                md = hashlib.md5(open(fthumb,'rb').read()).hexdigest()
                if md == m:
                    self.wp.cell(row=n, column=self.cpAlbumFile+1).value = 'AlbumImage'
                dest = self.copyFiles(r, fileOrig, fileName)
                fileLoc = dest.replace(self.treedir, "")
                self.wp.cell(row=n, column=self.cpDest+1).value = fileLoc
        print(children, 'bilder i', r[self.caItem], '-', r[self.caDescription])

    def copyFiles(self, row, origFile, fileName):
        subtreepath = ''
        for n in range (self.path_depth + 1):
            subtreepath = subtreepath + self.path[n]
            if n>0:
                subtreepath = subtreepath + '/'
        dest = subtreepath + fileName
        try:
            shutil.copy(origFile, dest)
        except Exception as e:
            #print ('FAILED COPY ', origFile, ' TO ', dest, ' ERROR MSG: ', e)
            with open(self.subdir + 'x-port.log') as logfile:
                logfile.write('FAILED COPY ', origFile, ' TO ', dest, ' >>> ERROR MSG: ', e, '\n')
        return dest

    def norskeBokstaver(self, name):
        
        name = name.replace("&#230;", "æ")
        name = name.replace("&#248;", "ø")
        name = name.replace('&#229;', "å")
        name = name.replace("&#198;", "Æ")
        name = name.replace("&#216;", "Ø")
        name = name.replace("&#197;", "Å")
        
        return name

    def cleanPathName(self, name):
        """Substitue some special characters"""
        name = self.norskeBokstaver(name)
        for a,b in self.custom_substitutions:
            name = name.replace(a,b)
        name = re.sub(r'[^a-zA-Z0-9]', '_', name)
        
        #name = name.decode('iso-8859-1')
        return name
    
    def getAlbumImages(self, p):
            
        for row in sheet.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if row[self.caParentDoc] == self.topParent:
                m = 2
        
    
a = xport()