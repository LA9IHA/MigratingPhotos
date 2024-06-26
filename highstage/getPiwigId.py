import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib
import re

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook

from cols import colsHighStage

# Purpose: Prepare HighStage album with photos for ingestion to Piwigo
# Pre requisites: Album.xlsx and Pic.xlsx is created from Highstage
# Licence: GNU 2.0
# Author: Ottar Kvindesland, 2024
# Reference: https://piwigo.miraheze.org/wiki/HighstageExport

class getId(colsHighStage):
    def __init__(self):
        
        super().__init__()
        
        pwgfile = self.injectdir + self.PiwigoPic
        self.pwg_wb = load_workbook(filename=pwgfile)
        self.wpwg = self.pwg_wb.worksheets[0]
        
        pwgafile = self.injectdir + self.PiwigoAlbum
        self.pwa_wb = load_workbook(filename=pwgafile)
        self.wpwa = self.pwa_wb.worksheets[0]
        
        picfile = self.subdir + self.fInputPic
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        albumfile = self.subdir + self.fInputAlbum
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        
        self.getSourcePicsId()
        self.getSourceAlbumId()
        
        self.pic_wb.save(self.subdir + self.fOutputPic)
        self.album_wb.save(self.subdir + self.fOutputAlbum)
        
    def getSourcePicsId(self):
        
        dst = self.cpDest+1
        pwgid = self.cpPiwigoId+1
        n = 0
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n += 1
            destpath = self.wp.cell(row=n, column=dst).value
            if destpath is not None and destpath != '' and n>1:
                pwgref = self.getPiwigoPicRef(destpath)
                self.wp.cell(row=n, column=pwgid).value = pwgref
    
    def getPiwigoPicRef(self, src):
        
        pwpath = self.pppath+1
        pwgid = self.ppid+1
        p_id = ''
        n = 0
        for pic in self.wpwg.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.pplastcol, values_only=False):
            n += 1
            dst = self.wpwg.cell(row=n, column=pwpath).value
            if dst is not None and dst.endswith(src):
                if p_id != '':
                    print ('ERROR, Pic: source file used multiple times: ', src)
                p_id = self.wpwg.cell(row=n, column=pwgid).value
        return p_id

    def getSourceAlbumId(self):
        
        dst = self.caAlbumPath+1
        pwgid = self.caPiwigoId+1
        n = 0
        for album in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n += 1
            destpath = self.wa.cell(row=n, column=dst).value
            if destpath is not None and destpath != '' and n>1:
                pwgref = self.getPiwigoAlbRef(destpath)
                self.wa.cell(row=n, column=pwgid).value = pwgref

    def getPiwigoAlbRef(self, src):
        
        pwpath = self.padir+1
        pwgid = self.paId+1
        n = 0
        p_id = ''
        if src.endswith('/'):
            src = src[:-1]
        for album in self.wpwa.iter_rows(min_row=1, max_row=self.wpwa.max_row, min_col=1, max_col=self.palastcol, values_only=False):
            n += 1
            dst = self.wpwa.cell(row=n, column=pwpath).value
            if dst is not None and src.endswith(dst):
                if p_id != '':
                    print ('ERROR, Album: source file used multiple times: ', src)
                p_id = str(self.wpwa.cell(row=n, column=pwgid).value)
        return p_id

a = getId()
