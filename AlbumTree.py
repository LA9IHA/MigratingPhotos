import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook

class AlbumTree:

    def __init__(self, h, t, s):
        self.homedir = h
        self.treedir = t
        self.subdir = s
        self.path_depth = 0
        
        #self.topParent = 'SPC1064-1A'
        self.topParent = 'ALBUM1054-1A'
                
        if os.path.exists(self.treedir):
            shutil.rmtree(self.treedir)
        os.makedirs(self.treedir)
        
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        self.path = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''] 
        self.path[0] = self.treedir
        self.MDerr = 'UFFDA'
        
        self.custom_image_extensions = ['.jpeg', '.png', '.jpg', '.gif']
        self.custom_substitutions = [('&', 'et'), ('(',''), (')',''), (' !', ''), ("'", ' '), (' ', '_')]
        self.custom_video_mime_types = ['media', 'video']
        
        albumfile = self.subdir + 'Album.xlsx'
        self.album_wb = load_workbook(filename=albumfile)
        self.wa = self.album_wb.worksheets[0]
        self.ca = 0
        self.caItem = 1
        self.caDescription = 2
        self.caWorkspace = 3
        self.caEventTime = 4
        self.caEditBy = 5
        self.caNote = 6
        self.caInitdate = 7
        self.caParentDoc = 8
        self.caFileName = 9
        self.caBareItem = 10
        self.caLastAlbum = 11
        self.caMD5 = 12
        
        picfile = self.subdir + 'Pic.xlsx'
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        self.cp = 0
        self.cpItem = 1
        self.cpDescription = 2
        self.cpWorkspace = 3
        self.cpEventTime = 4
        self.cpEditBy = 5
        self.cpNote = 6
        self.cpAlias = 7
        self.cpNote2 = 8 # Historikk
        self.cpDate2 = 9  # Første gangs arkivering
        self.cpDate3 = 10
        self.cpExif = 11
        self.cpInitdate = 12 # Date taken
        self.cpParentDoc = 13
        self.cpFileName = 14
        self.cpBareItem = 15
        self.cpLastPic = 16
        self.cpAlbumFile = 17
        
        self.makeTopAlbums(self.wa)
        self.album_wb.save(self.subdir + 'Album1.xlsx')
        self.pic_wb.save(self.subdir + 'Pic1.xlsx')
        
    def makeTopAlbums (self, sheet):
        for row in sheet.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if row[self.caParentDoc] == self.topParent:
                self.createTree(row)
                
    def createTree(self, row):
        self.path_depth += 1
        self.path[self.path_depth] = row[self.caDescription]
        md = self.createAMD5(row)
        self.createDir(row)
        self.createPics(row, md)
        self.createChildren(row, md)
        self.path_depth -= 1
        
    def createAMD5(self, r):
        md = self.MDerr
        if (r[self.caFileName] != ''):
            rnum = int(r[self.ca]) + 1
            fdir = self.subdir + 'ALBUMS/' + r[self.caBareItem] + '/' + r[self.caBareItem] + '/' + r[self.caFileName]
            md = hashlib.md5(open(fdir,'rb').read()).hexdigest()
            self.wa.cell(row=rnum, column=self.caMD5).value = md
        return md
        
    def createChildren(self, r, mdParent):
        for row in self.wa.iter_rows(min_row=1, max_row=self.wa.max_row, min_col=1, max_col=self.caLastAlbum, values_only=True):
            if row[self.caParentDoc] == r[self.caItem]:
                albFile = ''
                md = self.createAMD5(row)
                if md == mdParent:
                    albFile = 'Y'
                print ('CHILD ALBUM: ', row[self.caItem], ' - ', albFile)
                self.createTree(row)
       
        
    def createDir(self, row):
        subtreepath = ''
        for n in range (self.path_depth + 1):
            subtreepath = subtreepath + self.path[n]
            if n>0:
                subtreepath = subtreepath + '/'
        os.makedirs(self.cleanPathName(subtreepath))

    def createPics(self, r, m):
        n = 0
        children = 0
        c = self.cpParentDoc+1
        bi = self.cpBareItem+1
        fn = self.cpFileName+1
        ds = self.caDescription+1
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n+=1
            parent = self.wp.cell(row=n, column=c).value
            bareItem = self.wp.cell(row=n, column=bi).value
            fileName = self.wp.cell(row=n, column=fn).value
            if parent == r[self.caItem]:
                children += 1
                fpath = self.subdir + 'PHOTOS/' + bareItem + '/' + bareItem + '/'
                fthumb = fpath + 'doc_pic.jpg'
                ffull = fpath + fileName
                md = hashlib.md5(open(fthumb,'rb').read()).hexdigest()
                if md == m:
                    self.wp.cell(row=n, column=self.cpAlbumFile).value = 'AlbumImage'
                self.copyFiles(r, ffull, fileName)
        print(children, 'bilder i', r[self.caItem], '-', r[self.caDescription])

    def copyFiles(self, row, origFile, fileName):
        subtreepath = ''
        for n in range (self.path_depth + 1):
            subtreepath = subtreepath + self.path[n]
            if n>0:
                subtreepath = subtreepath + '/'
        dest = self.cleanPathName(subtreepath + fileName)
        shutil.copy(origFile, dest)

    def norskeBokstaver(self, name):
        
        name = name.replace("&#230;", "æ")
        name = name.replace("&#248;", "ø")
        name = name.replace('&#229;', "å")
        name = name.replace("&#198;", "Æ")
        name = name.replace("&#216;", "Ø")
        name = name.replace("&#197;", "Å")
        
        return name

    def cleanPathName(self, name):
        """Substitue some special characters"""
        name = self.norskeBokstaver(name)
        for a,b in self.custom_substitutions:
            name = name.replace(a,b)
        #name = name.decode('iso-8859-1')
        return name

homedir = "/Volumes/home/Oppgaver/billedfiler/" 
treedir = homedir + "tree/"
subdir = homedir + "highstage/"
    
a = AlbumTree(homedir, treedir, subdir)