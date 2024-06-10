import datetime
from openpyxl import Workbook
import openpyxl
import os
import shutil
import hashlib

from openpyxl.descriptors.base import DateTime
from openpyxl.reader.excel import load_workbook
from os import listdir

# Purpose: Prepare HighStage album with photos for  igration to Piwigo
# Pre requisites: Album.xlsx and Pic.xlsx is created from Highstage
# Licence: GNU 2.0
# Author: Ottar Kvindesland, 2024
# Reference: https://piwigo.miraheze.org/wiki/HighstageExport

class CheckPics:

    def __init__(self, h, t, s):
        self.homedir = h
        self.treedir = t
        self.subdir = s
        self.path_depth = 0
        
        self.creationDate = datetime.datetime(1980, 1, 1, 1, 0)
        self.path = ['', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', '', ''] 
        self.path[0] = self.treedir
        self.MDerr = 'UFFDA'
        
        self.custom_image_extensions = ['.jpeg', '.png', '.jpg', '.gif']
        self.custom_substitutions = [('&', 'et'), ('(',''), (')',''), (' !', ''), ("'", ' '), (' ', '_')]
        self.custom_video_mime_types = ['media', 'video']
        
        picfile = self.subdir + 'Pic.xlsx'
        self.pic_wb = load_workbook(filename=picfile)
        self.wp = self.pic_wb.worksheets[0]
        
        # Columns used in Pic.xlsx
        self.cp = 0 # A = ID
        self.cpItem = 1 # B = Item
        self.cpDescription = 2 # etc ...
        self.cpWorkspace = 3
        self.cpEventTime = 4
        self.cpEditBy = 5
        self.cpNote = 6
        self.cpAlias = 7
        self.cpNote2 = 8 # History
        self.cpDate2 = 9  # First time storage
        self.cpDate3 = 10
        self.cpExif = 11
        self.cpInitdate = 12 # Date taken
        self.cpParentDoc = 13
        self.cpFileName = 14
        self.cpBareItem = 15
        self.cpLastPic = 16
        self.cpAlbumFile = 17
        self.cpFileError = 18 # R - Error message
        
        self.checkPics()
        self.pic_wb.save(self.subdir + 'Pic1.xlsx')
        
    def checkPics(self):
        n = 0
        c = self.cpParentDoc+1
        bi = self.cpBareItem+1
        fn = self.cpFileName+1
        err = self.cpFileError+1
        for pic in self.wp.iter_rows(min_row=1, max_row=self.wp.max_row, min_col=1, max_col=self.cpLastPic, values_only=False):
            n+=1
            parent = self.wp.cell(row=n, column=c).value
            bareItem = self.wp.cell(row=n, column=bi).value
            fileName = self.wp.cell(row=n, column=fn).value
            fpath = self.subdir + 'PHOTOS/' + bareItem + '/' + bareItem + '/'
            fthumb = fpath + 'doc_pic.jpg'
            ffull = fpath + fileName
            
            if (bareItem.lower() != 'name'):
                if not(self.testFile(fpath)):
                    self.wp.cell(row=n, column=self.cpFileError).value = 'File Error'
                    print('ERROR: ' + fpath)
                
    def testFile(self, p):
        dp = 0
        smallFiles = 0
        goodSuffixes = 0
        
        for fil in os.listdir(p):
            ppath = os.path.join(p, fil)
            if os.path.isfile(ppath):
                fsize = os.path.getsize(ppath)
                
                if ppath.endswith('doc_pic.jpg'):
                    dp = dp + 1
                    if fsize < 750:
                        smallFiles = smallFiles + 1
                else:
                    if fsize < 1500:
                        smallFiles = smallFiles + 1
                    if (any(ppath.lower().endswith(filetype) for filetype in self.custom_image_extensions) == True):
                        goodSuffixes = goodSuffixes + 1
            
        #print('GS: ', goodSuffixes)
        return (( dp == 1) and (smallFiles == 0)  and (goodSuffixes == 1) )

homedir = "/Volumes/home/Oppgaver/hstransfer/" 
treedir = homedir + "tree/"
subdir = homedir + "source/"
    
a = CheckPics(homedir, treedir, subdir)
